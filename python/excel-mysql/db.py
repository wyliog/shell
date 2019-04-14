import pymysql.cursors
from openpyxl import load_workbook
import os
import sys

# 连接数据库
connection = pymysql.connect(host='192.168.142.130',  # 数据库地址
                             user='root',  # 数据库用户名
                             password='password',  # 数据库密码
                             db='excel',  # 数据库名称
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()


def excelToMySql(files, tablename):
    try:
    # 创建表
        sql = "CREATE TABLE IF NOT EXISTS `" + tablename + "` ( `id` INT UNSIGNED NOT NULL PRIMARY KEY AUTO_INCREMENT,`name` char(24) DEFAULT NULL," \
                                                           "`s1` char(24) DEFAULT NULL,`s2` varchar(24) DEFAULT NULL,`s3` char(24) DEFAULT NULL," \
                                                           "`s4` char(255) DEFAULT NULL,`ex_name` char(255) DEFAULT NULL) ENGINE=MyISAM DEFAULT CHARSET=utf8"
        # cursor.execute("DROP TABLE IF  EXISTS `"+tablename+"`")
        cursor.execute(sql)
        connection.commit()
        count = 0

        ex_name_sql = "select ex_name from %s " % tablename
        cursor.execute(ex_name_sql)
        ex_names = cursor.fetchall()
        for file in files:
            for ex_name in ex_names:
                if file in ex_name['ex_name']:
                    print(file, "数据已存在")
                    sys.exit(0)
            if os.path.splitext(file)[1] == '.xlsx':  # 如果是excel文件
                wb = load_workbook(file)
                table = wb["Sheet1"]
                print("导入", file, "数据中,已经导入", count, "条数据")
                excel_sql = 'insert into ' + tablename + ' values(%s,%s,%s,%s,%s,%s,%s)'
                for row in range(2, table.max_row + 1):
                    value = [None,
                             table.cell(row=row, column=1).value,
                             table.cell(row=row, column=2).value,
                             table.cell(row=row, column=3).value,
                             table.cell(row=row, column=4).value,
                             table.cell(row=row, column=5).value,
                             file]
                    print(value)
                    cursor.execute(excel_sql, value)
                    count += 1
                connection.commit()

        print('批量导入数据完毕，共导入:', count, '条数据')
    finally:
        # 关闭数据库
        connection.close()


if __name__ == '__main__':
    basedir = 'E:\code\pythondemo'
    files = os.listdir(basedir)
    print(files)
    excelToMySql(files, 'test1')
